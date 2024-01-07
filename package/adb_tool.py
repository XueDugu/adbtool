import datetime
import os
import threading
import time
import tkinter as tk
import subprocess
from openpyxl import Workbook, load_workbook

custom_functions = {
    "log": lambda: None,
    "back": lambda: None,
    "reverse": lambda: None,
    "activity": lambda: None,
    "burpopen": lambda: None,
    "burpclose": lambda: None,
    "unlock": lambda: None,
}

def run_adb_command(command, success_message):
    command_function = globals().get(command)
    if command_function:
        command_function()
        text_box.insert(tk.END, success_message + "\n")
    else:
        text_box.insert(tk.END, f"Error: Function {command} not found!\n")

def run_custom_function(command, success_message):
    custom_function = custom_functions.get(command)
    if custom_function:
        custom_function()
        text_box.insert(tk.END, success_message + "\n")
    else:
        text_box.insert(tk.END, f"Error: Function {command} not found!\n")

def create_button(root, text, adb_command):
    btn = tk.Button(root, text=text, command=lambda c=adb_command, t=text: run_custom_function(c, f"{t} success!"), **button_style)
    return btn

def get_button_style():
    return {"width": 15, "height": 2, "padx": 50, "pady": 10, "bd": 4, "relief": "raised", "bg": "lightblue", "fg": "black", "font": ('Helvetica', 12, 'bold')}

root = tk.Tk()
root.title("adb工具")

text_label = tk.Label(root, text="记录栏", font=('Helvetica', 12, 'bold'))
text_label.grid(row=0, column=2, columnspan=2, pady=10)

text_box = tk.Text(root, height=15, width=50)
text_box.grid(row=1, column=2, rowspan=3, columnspan=2, padx=10, pady=10)

def clear_text_box():
    text_box.delete("1.0", tk.END)

adb_commands = {
    "Log": "log",
    "Back": "back",
    "Reverse": "reverse",
    "Monitor": "activity",
    "Burp Open": "burpopen",
    "Burp Close": "burpclose",
    "Unlock": "unlock",
}
button_style = get_button_style()

# 使用栈来记录用户点击的按钮
button_stack = []
run_stack = []

def repeat_last_command(repeat_count_entry):
    try:
        repeat_count = int(repeat_count_entry.get())
        repeat_count = max(1, repeat_count)
    except ValueError:
        repeat_count = 1

    # 从栈中取出最近的按钮进行重复操作
    run_stack.clear()
    for _ in range(repeat_count):
        if button_stack:
            button_text = button_stack.pop()
            run_stack.append(button_text)

    # 从栈中取出最近的按钮进行重复操作
    for _ in range(repeat_count):
        if run_stack:
            run_text = run_stack.pop()
            button_stack.append(run_text)
            last_command = adb_commands.get(run_text)
            if last_command:
                run_adb_command(last_command, f"Repeated {run_text} success!")

def show_previous_commands():
    text_box.insert(tk.END, "Previous Commands: " + " -> ".join(button_stack) + "\n")

def delete_stack_commands(delete_count_entry):
    try:
        delete_count = int(delete_count_entry.get())
        delete_count = max(1, delete_count)
    except ValueError:
        delete_count = 1

    for _ in range(delete_count):
        text_box.insert(tk.END, "Success delete " + button_stack.pop() + " !\n")

    update_label()

for i, (button_text, adb_command) in enumerate(adb_commands.items()):
    btn = create_button(root, button_text, adb_command)
    btn.grid(row=i // 2 + 1, column=i % 2)
    btn.bind("<Button-1>", lambda event, btn=button_text: set_last_buttons_clicked(btn))

clear_btn = tk.Button(root, text="Clear", command=clear_text_box, **button_style)
clear_btn.grid(row=len(adb_commands) // 2 + 2, column=3, pady=10)

repeat_count_entry = tk.Entry(root)
repeat_count_entry.grid(row=len(adb_commands) // 2 + 3, column=2, pady=10)

repeat_btn = tk.Button(root, text="Repeat", command=lambda: repeat_last_command(repeat_count_entry), **button_style)
repeat_btn.grid(row=len(adb_commands) // 2 + 3, column=3, pady=10)

previous_btn = tk.Button(root, text="Use History", command=show_previous_commands, **button_style)
previous_btn.grid(row=len(adb_commands) // 2 + 4, column=3, pady=10)

delete_count_entry = tk.Entry(root)
delete_count_entry.grid(row=len(adb_commands) // 2 + 5, column=2, pady=10)
delete_btn = tk.Button(root, text="Delete Stack", command=lambda: delete_stack_commands(delete_count_entry), **button_style)
delete_btn.grid(row=len(adb_commands) // 2 + 5, column=3, pady=10)

def set_last_buttons_clicked(button_text):
    # 将点击的按钮入栈
    button_stack.append(button_text)
    update_label()

size_label = tk.Label(root, text="Size of Stack:" + str(len(button_stack)), font=('Helvetica', 12, 'bold'))
size_label.grid(row=len(adb_commands) // 2 + 1, column=2, columnspan=2, pady=10)

def update_label():
    # 使用当前的button_stack长度更新标签文本
    size_label.config(text="Size of Stack:" + str(len(button_stack)))

    # 根据button_stack的长度更新标签的位置
    size_label.grid(row=len(adb_commands) // 2 + 1, column=2, columnspan=2, pady=10)

root.mainloop()

def run_adb_command_activity(command):
    try:
        adb_command = f"adb shell {command}"
        output = subprocess.check_output(adb_command, shell=True)
        with open('adb_output.txt', 'w') as f:
            f.write(output.decode('utf-8'))
        print("命令执行成功！")
    except subprocess.CalledProcessError as e:
        print(f"命令执行失败：{e}")

# 执行adb命令
def activity():
    print("All:\n")
    command = 'su -c "dumpsys window"'
    run_adb_command_activity(command)
    print("\nOn Windows:\n")
    command = 'su -c "dumpsys window | grep Current"'
custom_functions["activity"] = activity 
def run_adb_command_back(command):
    try:
        adb_command = f"adb shell {command}"
        subprocess.check_output(adb_command, shell=True)
        # with open('adb_output.txt', 'a') as f:
        #     f.write(output.decode('utf-8') + "\n")
        print("命令执行成功！")
    except subprocess.CalledProcessError as e:
        print(f"命令执行失败：{e}")
        
def back():
    command='su -c "input keyevent 4"'
    run_adb_command_back(command)
custom_functions["back"] = back 

def run_adb_command_burpclose(command):
    try:
        result = subprocess.run(command, shell=True, check=True, text=True, capture_output=True)
        if result.returncode == 0:
            print("操作成功！")
        else:
            print("操作失败！")
            print(result.stderr)
    except subprocess.CalledProcessError as e:
        print(f"命令执行失败：{e}")
    except Exception as e:
        print(f"发生未知错误：{e}")

def burpclose():
# 执行adb命令
    package_name = "com.github.kr328.clash"

# 关闭应用
    close_command = f'adb shell su -c "am force-stop {package_name}"'
    run_adb_command_burpclose(close_command)
custom_functions["burpclose"] = burpclose
def run_adb_command_burpopen(command):
    try:
        result = subprocess.run(command, shell=True, check=True, text=True, capture_output=True)
        if result.returncode == 0:
            print("操作成功！")
        else:
            print("操作失败！")
            print(result.stderr)
    except subprocess.CalledProcessError as e:
        print(f"命令执行失败：{e}")
    except Exception as e:
        print(f"发生未知错误：{e}")

def burpopen():
# 执行adb命令
    package_name = "com.github.kr328.clash"
    activity_name = "com.github.kr328.clash.MainActivity"
    command = f'adb shell su -c "am start -n {package_name}/{activity_name}"'
    run_adb_command_burpopen(command)
    time.sleep(0.5)
    command_start=f'adb shell su -c "input tap 690 690"'
    run_adb_command_burpopen(command_start)
custom_functions["burpopen"] = burpopen

def log():
    file_name = 'log.xlsx'
    times=[0,0,0,0,0]

    if not os.path.isfile(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["Time", "Error", "Warning", "Info", "Debug", "Verbose"])
        wb.save(file_name)
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    def log_to_excel(log_level, message):
        if log_level == "E":
            times[0]+=1
            if times[0]<=5:
                row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), message, "", "", "", ""]
                ws.append(row)
        elif log_level == "W":
            times[1]+=1
            if times[1]<=5:
                row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", message, "", "", ""]
                ws.append(row)
        elif log_level == "I":
            times[2]+=1
            if times[2]<=5:
                row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", message, "", ""]
                ws.append(row)
        elif log_level == "D":
            times[3]+=1
            if times[3]<=5:
                row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", message, ""]
                ws.append(row)
        elif log_level == "V":
            times[4]+=1
            if times[4]<=5:
                row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", "", message]
                ws.append(row)
        wb.save(file_name)

    def run_adb_command(command, log_level):
        try:
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, text=True)
            output = process.communicate()[0]
            if output:
                print(output.strip())
                log_to_excel(log_level, output.strip())

            return_code = process.returncode
            if return_code != 0:
                print(f"命令执行失败，返回代码：{return_code}")
        except subprocess.CalledProcessError as e:
            print(f"命令执行失败：{e}")


# 运行10秒
    timeout = time.time() + 10
    while time.time() < timeout:
        log_levels = ["E", "W", "I", "D", "V"]
        for level in log_levels:
            t = threading.Thread(target=run_adb_command, args=(f'adb logcat *:{level} -t 1', level))
            t.start()
        time.sleep(1)

# 关闭前保存一次工作簿
    wb.save(file_name)
custom_functions["log"] = log
def run_adb_command_reverse(command):
    try:
        result = subprocess.run(command, shell=True, check=True, text=True, capture_output=True)
        if result.returncode == 0 or result.stdout.strip() == "8080":
            print("操作成功！")
            if result.stdout.strip() == "8080":
                print("返回值为8080")
            else:
                print("无返回值")
        else:
            print("操作失败！")
            print(result.stderr)
    except subprocess.CalledProcessError as e:
        print(f"命令执行失败：{e}")

def reverse():
# 执行adb命令
    command = 'adb reverse tcp:8080 tcp:8080'
    run_adb_command_reverse(command)
custom_functions["reverse"] = reverse
def run_adb_command_unlock(command):
    try:
        result = subprocess.run(command, shell=True, check=True, text=True, capture_output=True)
        if result.returncode == 0:
            print("操作成功！")
        else:
            print("操作失败！")
            print(result.stderr)
    except subprocess.CalledProcessError as e:
        print(f"命令执行失败：{e}")
    except Exception as e:
        print(f"发生未知错误：{e}")

def unlock():
# 亮屏
    close_command = f'adb shell su -c "input keyevent 82"'
    run_adb_command_unlock(close_command)
    # 显示解锁界面
    run_adb_command_unlock(close_command)
    time.sleep(0.2)
    # 输入密码
    command_0=f'adb shell su -c "input tap 780 2380"'
    run_adb_command_unlock(command_0)
    run_adb_command_unlock(command_0)
    run_adb_command_unlock(command_0)
    run_adb_command_unlock(command_0)
    time.sleep(0.1)
    # 解锁进入
    command_go=f'adb shell su -c "input tap 1140 2390"'
    run_adb_command_unlock(command_go)
    # 保持亮屏1分钟
    command_light=f'adb shell su -c "settings put system screen_off_timeout 60000"'
    run_adb_command_unlock(command_light)

custom_functions["unlock"] = unlock