import os
import subprocess
import threading
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

file_name = 'log.xlsx'
times=[0,0,0,0,0]

if not os.path.isfile(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Time", "Error", "Warning", "Info", "Debug", "Verbose"])
    wb.save(file_name)
else:
    wb = load_workbook(file_name)
    ws = wb.active

def log_to_excel(log_level, message):
    if log_level == "E":
        times[0]+=1
        if times[0]<=5:
            row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), message, "", "", "", ""]
            ws.append(row)
    elif log_level == "W":
        times[1]+=1
        if times[1]<=5:
            row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", message, "", "", ""]
            ws.append(row)
    elif log_level == "I":
        times[2]+=1
        if times[2]<=5:
            row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", message, "", ""]
            ws.append(row)
    elif log_level == "D":
        times[3]+=1
        if times[3]<=5:
            row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", message, ""]
            ws.append(row)
    elif log_level == "V":
        times[4]+=1
        if times[4]<=5:
            row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", "", message]
            ws.append(row)
    wb.save(file_name)

def run_adb_command(command, log_level):
    try:
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, text=True)
        output = process.communicate()[0]
        if output:
            print(output.strip())
            log_to_excel(log_level, output.strip())

        return_code = process.returncode
        if return_code != 0:
            print(f"命令执行失败，返回代码：{return_code}")
    except subprocess.CalledProcessError as e:
        print(f"命令执行失败：{e}")

# 运行10秒
timeout = time.time() + 10
while time.time() < timeout:
    log_levels = ["E", "W", "I", "D", "V"]
    for level in log_levels:
        t = threading.Thread(target=run_adb_command, args=(f'adb logcat *:{level} -t 1', level))
        t.start()
    time.sleep(1)

# 关闭前保存一次工作簿
wb.save(file_name)
